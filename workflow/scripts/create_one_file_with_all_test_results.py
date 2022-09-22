import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


bool_df = pd.read_csv(snakemake.input[0], sep='\t', dtype='object', index_col=[0, 1])
scoring_df = pd.read_csv(snakemake.input[1], sep='\t', dtype='object', index_col=[0, 1])

with pd.ExcelWriter(snakemake.output[0], engine='openpyxl') as writer:
    bool_df.to_excel(writer, sheet_name="results", startrow=1)
    scoring_df.to_excel(writer, sheet_name="results", startrow=bool_df.shape[0] + 5)


# styling
h1_font = Font(color="FF0000")
h2_font = Font(bold=True)
h1_fill = PatternFill("solid", fgColor="92D050")


def remove_all_style(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.style = 'Normal'


wb = load_workbook(snakemake.output[0])
ws = wb['results']
remove_all_style(ws)

# column header
for cell in ws['2'] + ws['{}'.format(bool_df.shape[0] + 5 + 1)]:
    cell.font = h2_font

# row header
for cell in ws['A']:
    if isinstance(cell, openpyxl.cell.cell.Cell):
        cell.alignment = Alignment(
            horizontal='general',
            vertical='top'
        )

# category header
ws['A1'] = "For binary_not_depleted"
ws['A1'].font = h1_font
ws['A1'].fill = h1_fill

ws['A{}'.format(bool_df.shape[0] + 5)] = "For scoring_not_depleted"
ws['A{}'.format(bool_df.shape[0] + 5)].font = h1_font
ws['A{}'.format(bool_df.shape[0] + 5)].fill = h1_fill


# cell width
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 25

for i in range(3, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(i)].width = 15

wb.save(snakemake.output[0])
